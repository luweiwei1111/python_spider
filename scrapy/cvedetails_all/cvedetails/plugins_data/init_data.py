# -*- coding: utf-8 -*-
from cvedetails.sqlitepiplines.sql import Sql
from cvedetails import settings 
import openpyxl
import xlrd
import xlwt
import datetime
import os
import re
import sys
import codecs

class ExcelRead:
    def __init__(self):
        self.excelSettingsFileName = settings.ExcelSettingsFileName
        Sql.ctl_tb_settings()
        Sql.ctr_tb_settings()

    def read_excel_to_db(self):
        wb = openpyxl.load_workbook(self.excelSettingsFileName)

        # 获取workbook中所有的表格
        sheets = wb.sheetnames        
        print(sheets)

        # 循环遍历所有sheet
        for i in range(len(sheets)):
            if i == 0:
                continue

            sheet = wb[sheets[i]]
            print('\n\n第' + str(i) + '个sheet: ' + sheet.title + '->>>')
            sheet_title = sheet.title
            for r in range(1, sheet.max_row + 1):
                category = ''
                product_name = ''
                product_id = ''
                vendor_id = ''
                product_search = ''
                vendor_search = ''
                owner = ''
                if r == 1:
                    print('\n' + ''.join(
                        [str(sheet.cell(row=r, column=c).value).ljust(17) for c in range(1, sheet.max_column + 1)]))
                else:
                    category = str(sheet.cell(row=r, column=1).value)
                    product_name = str(sheet.cell(row=r, column=2).value)
                    product_id = str(sheet.cell(row=r, column=3).value)
                    vendor_id = str(sheet.cell(row=r, column=4).value)
                    product_search = str(sheet.cell(row=r, column=5).value)
                    vendor_search = str(sheet.cell(row=r, column=6).value)
                    owner = str(sheet.cell(row=r, column=7).value)
                    #print('category=%s, product_name=%s, product_id=%s, vendor_id=%s, product_search=%s, vendor_search=%s, owner=%s' % (category, product_name, product_id, vendor_id, product_search, vendor_search, owner))
                    Sql.insert_tb_settings(category, product_name, product_id, vendor_id, product_search, vendor_search, owner)


class excelProcess:
    def __init__(self):
        self.excelFileName = settings.ExcelReportFileName    #excel文件名
        self.excelSheetName = settings.ExcelReportSheetName  #excel sheet名

    def WriteLog(self, message,date):
        fileName = os.path.join(os.getcwd(),  date  +   '.txt')
        with open(fileName, 'a') as f:
            f.write(message)

    def WriteSheetRow(self,sheet, rowValueList, rowIndex, isBold):
        i = 0
        style = xlwt.easyxf('font: bold 1')
        # style = xlwt.easyxf('font: bold 0, color red;')#红色字体
        # style2 = xlwt.easyxf('pattern: pattern solid, fore_colour yellow; font: bold on;') # 设置Excel单元格的背景色为黄色，字体为粗体
        for svalue in rowValueList:
            if isBold:
                sheet.write(rowIndex, i, svalue, style)
            else:
                sheet.write(rowIndex, i, svalue)
            i = i + 1

    def save_Excel(self):
        wbk = xlwt.Workbook()
        sheet = wbk.add_sheet(self.excelSheetName, cell_overwrite_ok=True)       

        headList = ['product_id', 'product_name', 'year', 'vul_type', 'cve','openvas_file',
                    'openvas_exist', 'topvas_ness_file', 'nessus_file', 'nessus_exist','ts_file', 'ts_count']

        rowIndex = 0
        self.WriteSheetRow(sheet, headList, rowIndex, True)
        reports_list = Sql.select_from_cve_report()
        for report_info in reports_list:
            rowIndex = rowIndex + 1
            valueList = []
            for i in range(0, len(report_info)):
                valueList.append(report_info[i])
                self.WriteSheetRow(sheet, valueList, rowIndex, False)
        fileName = os.path.join(os.getcwd(), self.excelFileName)
        wbk.save(fileName)

class PluginsWriteData:
    def __init__(self, plugin_type):
        print('##Begin write plugins data for %s' %(plugin_type))
        if plugin_type == 'topvas':
            self.plugins_path = settings.Topvas_PATH
            self.plugins_table = 'nvts'
            Sql.ctl_tb_plugins(self.plugins_table)
            Sql.cls_tb_plugins(self.plugins_table)
        elif plugin_type == 'nessus':
            self.plugins_path = settings.Nessus_PATH
            self.plugins_table = 'nvts_ness'
            Sql.ctl_tb_plugins(self.plugins_table)
            Sql.cls_tb_plugins(self.plugins_table)
        else:
            sys.exit('plugin_type[%s] error!!!' %(plugin_type))

        self.count = 0
        
        return

    def description_parse(self, line, pf, state):
        #检测注释
        line = line.lstrip()
        if  len(line) == 0 or '#' == line[0]:
            return state
    
        #查找 if (description) 或者 if (description) {
        if state == 1:
            ctx = re.search(r'(^|;)\s*if\s*\(description\)\s*$',line)
            if ctx:
                state = 2;
                return state
                
            ctx = re.search(r'(^|;)\s*if\s*\(description\)\s*{',line)
            if ctx:
                state = 3;
                line = line[ctx.span()[1]:]
    
        #已经查找到 if (description), 现在查找 {
        if state == 2:
             ctx = re.search(r'\s*{',line);
             if ctx:
                state = 3;
                line = line[ctx.span()[1]:]
    
        #已经查找到  if (description) {, 现在匹配相关数据，并且查找 }
        if state == 3:
            while 1:
                # if self.cves != None and self.plugin_id != None:
                #     print('cves:%s, plugin_id:%s' %(self.cves, self.plugin_id))
                #     state = 0
                #     break
                if re.search(r'^\s*}', line):
                    state = 0
                    break
                    
                end_line = self.tag_parse(line, pf)
    
                if end_line:
                    line = end_line
                    if re.search(r'^\s*}', end_line):
                        state = 0
                        break
                else:
                    break;
    
        return state
    
    #检测格式tag(ddd){}
    def tag_parse(self, line, pf):
        #这部分不能缺少
        line, end_line = self.get_full_tag(line, pf)
        
        if self.cves != None and self.plugin_id != None:
            return end_line
            
        line = line.strip()
        
        #判断空白行或者注释
        if len(line) <= 0 or line[0] == "#":
            return end_line

        if self.tag_check(line, "script_id") or self.tag_check(line, "script_oid"):
            line = line.replace("script_id",""); 
            line = line.replace("script_oid","");
            line = line.replace('"',"");
            line = line.replace("'","");
            line = line.replace('(',"");
            line = line.replace(")","");
            line = line.replace('\r',"");
            line = line.replace("\n","");
            self.plugin_id = line

        elif self.tag_check(line, "script_cve_id"):
            line = line.replace("script_cve_id","");
            line = line.replace('"',"");
            line = line.replace("'","");
            line = line.replace('(',"");
            line = line.replace(")","");
            line = line.replace('\r',"");
            line = line.replace("\n","");
            self.cves = line

        return end_line

    def tag_check(self, line, pattern):
        match = r'^\s*'+ pattern + r'\s*\('
        ctx = re.search(match, line)
        if ctx:
            return True
        return False

    def get_full_tag(self, line, pf):
        stack = ''
        tmpline=""
        endline = None

        while True:
            stack, newline, endline = self.line_check(line, pf, stack)
            tmpline += newline

            if len(stack) == 0:
                break;
            line = pf.readline()
            if not line:
                print(self.path + "   ------>parse faild!!!!!")
                break;
            
        return tmpline, endline
        
    def line_check(self, line, pf, stack):
        set = '(){}\"\'#;'
        tmpline = ""
        endline = None

        for i in range(len(line)):
            char = line[i]
            if len(stack) > 0:
                if stack[-1] == '#':
                    if char == '\n':
                        stack = stack[0:-1]
                    continue;
                elif stack[-1] == '"':
                    if char == '"':
                        if line[i-1] != '\\':
                            stack = stack[0:-1]
                    tmpline += char
                    continue;
                elif stack[-1] == "'":
                    if char == "'":
                        if line[i-1] != '\\':
                            stack = stack[0:-1]
                    tmpline += char
                    continue
                    
            if char in set:
                if char == '#':
                    stack += char
                    continue
                    
                if char == ')':
                    if stack[-1] != '(':
                        print(os.path.basename(pf.name) + ':parse error ) !!\n')
                    stack = stack[0:-1]
                elif char == '}':
                    if stack[-1] != '{':
                        print(os.path.basename(pf.name) + ':parse error } !!\n')
                    stack = stack[0:-1]
                elif char ==';':
                    if len(stack) == 0:
                        endline = line[i+1: len(line)]
                        break;
                else:
                    stack += char
                    
            tmpline += char
        return stack, tmpline, endline

    def db_inset(self, path, cves, id):
        cve_str = ""
        for cve in cves:
            cve_str = cve_str + cve.strip() + ","

        cve_str = cve_str[0:-1]
        Sql.insert_plugins(self.plugins_table, id, cve_str, path)      

    def nasl_parse(self, path):
        if 0 == self.count % 1000:
            print(self.count)

        self.count += 1
        self.path = path

        pf = codecs.open(path, 'r+', encoding = "ISO-8859-1")

        #初始化cves及id数据
        self.cves = None
        self.plugin_id = None
        
        #初始状态, 没有接续结束
        state = 1
        while True:
            line = pf.readline()
            if not line:
                break;
            if state > 0:
                state = self.description_parse(line, pf, state)
            else:
                break;
                
        #解析结束, 写数据库
        if 0 == state:
            result = re.match(".*/", path)
            path = path[result.span()[1]:len(path)]
            
            if self.cves is not None:
                if self.plugin_id is None:
                    self.plugin_id = "unknow"
                self.db_inset(path, self.cves.split(','), self.plugin_id)

        pf.close()

    def file_walk(self):
        rootdir = self.plugins_path
        #print('dir:' + rootdir)
        list = os.listdir(rootdir)
        for i in range(0,len(list)):
            path = os.path.join(rootdir,list[i])
            if os.path.isfile(path):
                if re.match(".*\.nasl$", path) is not None:
                    self.nasl_parse(path);
            elif os.path.isdir(path):
                self.plugins_path = path
                self.file_walk()

       
    def Del_XX_nasl(self):
        #删除XX_*.nasl文件
        Sql.delete_XX_nasl()

class TopVAS:
    def __init__(self):
        #Sql.drop_tb_cve_report()
        Sql.ctl_tb_cve_report()
        Sql.cls_tb_cve_report()
        Sql.ctl_index_nvts_ness()
        
        Sql.ctl_tb_nvts_nons()
        Sql.ctl_index_nvts_nons()
        Sql.insert_nvts_nons()

    def cve_report(self):
        #cve topvas
        cve_detail_list = Sql.select_cve_detail_list()
        
        progress = 0
        for cve_info in cve_detail_list:
            product_id = cve_info[0]
            product_name = cve_info[1]
            year = cve_info[2]
            vul_type = cve_info[3]
            cve = cve_info[4]

            #topvas
            nvt_topvas_list = Sql.select_nvts_topvas_by_cve(cve)
            openvas_exist = 'no'
            openvas_file = ''
            topvas_ness_file = ''
            if len(nvt_topvas_list) != 0:
                openvas_exist = 'yes'
                openvas_file = ''
                count = 0
                for file in nvt_topvas_list:
                    count = count + 1
                    if count == 1:
                        openvas_file = file[0]
                    else:
                        openvas_file = openvas_file + ',' + file[0]

            #nessus
            nvt_ness_list = Sql.select_nvts_ness_by_cve(cve)
            nessus_file = ''
            nessus_exist = 'no'
            topvas_ness_file_list = []
            if len(nvt_ness_list) != 0:
                nessus_exist = 'yes'
                nessus_file = ''
                count = 0
                for file in nvt_ness_list:
                    count = count + 1
                    ns_file =  'ns_' + file[0]
                    ret = Sql.select_nvts_by_file(ns_file)
                    if ret[0] == 1:
                        print('file:%s存在' + file[0])
                        topvas_ness_file_list.append(file[0])
                        topvas_ness_file = topvas_ness_file + ',' + ns_file
                    
                    nessus_file = nessus_file + ',' + file[0]
            
            if ',' in nessus_file:
                nessus_file = nessus_file[1:]

            nessus_list = nessus_file.split(',')
            ts_count = 0
            ts_file = ''
            if openvas_exist == 'no' and nessus_exist == 'yes':
                if len(topvas_ness_file_list) == 0:
                    ts_count = len(nessus_list)
                    ts_file = nessus_file
                else:
                    for ness_info in nessus_list:
                        if ness_info not in topvas_ness_file_list:
                            ts_count = ts_count + 1
                            ts_file = ts_file + ',' + ness_info
                    if ',' in ts_file:
                        ts_file = ts_file[1:]
            
            #去除首位逗号,
            if ',' in topvas_ness_file:
                topvas_ness_file = topvas_ness_file[1:]

            #生成报告
            progress = progress + 1
            #if progress == 101:
            #    return
            print('###progress:%d' %(progress))
            Sql.insert_cve_report(product_id, product_name, year, vul_type, cve, openvas_file, openvas_exist, topvas_ness_file, nessus_file, nessus_exist, ts_file, ts_count)

        #删除表nvts_nons
        Sql.drop_tb_nvts_nons()